"""Generates the multi-sheet analysis workbook from analysis_data.py."""

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

sys.path.insert(0, str(Path(__file__).parent))
import analysis_data as D  # noqa: E402

# --- Palette ---------------------------------------------------------------
INK = "1A1A1A"
MUTED = "6B6B6B"
RULE = "D8D4CE"
BAND = "F7F5F2"
ACCENT = "1F4E5F"
ACCENT_LIGHT = "E8F0F2"

OK = "1E7A46"
OK_BG = "E6F4EC"
WARN = "8A5A00"
WARN_BG = "FDF3E0"
BAD = "9B2C2C"
BAD_BG = "FBEAEA"
NEUTRAL_BG = "EFEFEF"

TITLE_F = Font(name="Calibri", size=18, bold=True, color=ACCENT)
SUB_F = Font(name="Calibri", size=11, color=MUTED)
HEAD_F = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BODY_F = Font(name="Calibri", size=10, color=INK)
BODY_SM = Font(name="Calibri", size=9, color=MUTED)
MONO_F = Font(name="Consolas", size=9, color=INK)
KPI_NUM = Font(name="Calibri", size=22, bold=True, color=ACCENT)

HEAD_FILL = PatternFill("solid", fgColor=ACCENT)
BAND_FILL = PatternFill("solid", fgColor=BAND)

THIN = Side(style="thin", color=RULE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def status_style(value: str):
    v = (value or "").strip().lower()
    if v in ("pass", "closed", "complete", "implemented", "full", "yes"):
        return Font(name="Calibri", size=10, bold=True, color=OK), PatternFill("solid", fgColor=OK_BG)
    if v.startswith("implemented") or v.startswith("p0 and p1") or v.startswith("phases a-c"):
        return Font(name="Calibri", size=10, bold=True, color=OK), PatternFill("solid", fgColor=OK_BG)
    if v in ("partial", "managed", "next", "continuous", "by design", "accepted, monitored",
             "deferred, deliberately", "verified by measurement"):
        return Font(name="Calibri", size=10, bold=True, color=WARN), PatternFill("solid", fgColor=WARN_BG)
    if v in ("none", "open", "blocked", "fail", "no"):
        return Font(name="Calibri", size=10, bold=True, color=BAD), PatternFill("solid", fgColor=BAD_BG)
    if v in ("n/a", "out of scope", "backlog", "backlog, triggered", "backlog, conditional"):
        return Font(name="Calibri", size=10, color=MUTED), PatternFill("solid", fgColor=NEUTRAL_BG)
    return BODY_F, PatternFill("solid", fgColor=BAND)


def sheet_title(ws, title: str, subtitle: str, span: int) -> int:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=title)
    c.font = TITLE_F
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    c = ws.cell(row=2, column=2 - 1, value=subtitle)
    c.font = SUB_F
    c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28
    return 4


def write_table(ws, start_row: int, headers, rows, widths,
                status_cols=(), mono_cols=(), name="T"):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=i, value=h)
        c.font = HEAD_F
        c.fill = HEAD_FILL
        c.alignment = CENTER
        c.border = BOX
    ws.row_dimensions[start_row].height = 30

    for r, row in enumerate(rows, start=start_row + 1):
        banded = (r - start_row) % 2 == 0
        for i, value in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=value)
            c.border = BOX
            c.alignment = WRAP
            if i in status_cols:
                f, fill = status_style(str(value))
                c.font, c.fill = f, fill
                c.alignment = CENTER
            elif i in mono_cols:
                c.font = MONO_F
                if banded:
                    c.fill = BAND_FILL
            else:
                c.font = BODY_F
                if banded:
                    c.fill = BAND_FILL

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    end_row = start_row + len(rows)
    if len(rows) > 0:
        ref = f"A{start_row}:{get_column_letter(len(headers))}{end_row}"
        t = Table(displayName=name, ref=ref)
        t.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False)
        ws.add_table(t)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    return end_row


def note(ws, row: int, text: str, span: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = BODY_SM
    c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 30
    return row + 2


wb = Workbook()

# ===========================================================================
# Sheet 1 - Executive summary
# ===========================================================================
ws = wb.active
ws.title = "Summary"
ws.sheet_view.showGridLines = False

ws.merge_cells("A1:H1")
c = ws["A1"]
c.value = D.META["title"]
c.font = Font(name="Calibri", size=24, bold=True, color=ACCENT)
ws.row_dimensions[1].height = 36

ws.merge_cells("A2:H2")
c = ws["A2"]
c.value = D.META["subtitle"]
c.font = Font(name="Calibri", size=12, color=MUTED)
ws.row_dimensions[2].height = 22

ws.merge_cells("A3:H3")
c = ws["A3"]
c.value = f"Version {D.META['version']}   |   {D.META['date']}   |   Basis: {D.META['basis']}   |   Code review: {D.SHIPPED_META['package']}"
c.font = BODY_SM
ws.row_dimensions[3].height = 20

# KPI tiles
tiles = [
    ("1 / 10", "shipped evaluator agrees", "with a reviewer"),
    ("10 / 10", "V2 evaluator agrees", "same ten answers"),
    ("2", "Critical findings", "both fixes under 30 lines"),
    ("93", "tests passing", "0 failures"),
    ("0.356 ms", "evaluation p95", "of a 300 ms budget"),
    ("0", "LLM calls on any path", "fully deterministic"),
]
row = 5
for i, (num, label, sub) in enumerate(tiles):
    col = 1 + (i % 3) * 3
    r = row + (i // 3) * 4
    ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 1)
    cc = ws.cell(row=r, column=col, value=num)
    cc.font = KPI_NUM
    cc.alignment = Alignment(vertical="center")
    cc.fill = PatternFill("solid", fgColor=ACCENT_LIGHT)
    ws.row_dimensions[r].height = 34

    ws.merge_cells(start_row=r + 1, start_column=col, end_row=r + 1, end_column=col + 1)
    cc = ws.cell(row=r + 1, column=col, value=label)
    cc.font = Font(name="Calibri", size=10, bold=True, color=INK)
    cc.fill = PatternFill("solid", fgColor=ACCENT_LIGHT)

    ws.merge_cells(start_row=r + 2, start_column=col, end_row=r + 2, end_column=col + 1)
    cc = ws.cell(row=r + 2, column=col, value=sub)
    cc.font = BODY_SM
    cc.fill = PatternFill("solid", fgColor=ACCENT_LIGHT)

row = 14
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
c = ws.cell(row=row, column=1, value="What was asked, and what this answers")
c.font = Font(name="Calibri", size=13, bold=True, color=ACCENT)
row += 1

findings = [
    ("The shipped evaluator agrees with a reviewer once in ten answers, and three of the misses award marks the student did not earn.",
     "Key points are matched with a raw substring test. A student who writes 'there is no evidence of deep vein thrombosis' "
     "receives full credit for the DVT key point; 'inTESTinal' satisfies a key point of 'test'; a fully correct Arabic answer "
     "scores zero. The V2 evaluator agrees 10/10 on the same set. The swap is one file and changes no API."),
    ("A student can resubmit any question after seeing the score.",
     "Answers are written with INSERT OR REPLACE and no route checks for an existing answer, so every question is effectively "
     "unlimited attempts with full feedback between them - and the results table keeps only the last attempt. One statement fixes it."),
    ("The framework is sound; the gaps were in enforcement, not design.",
     "Every invariant it states as prose is now a compile-time type, a database constraint, or a test. The three that were only prose - no path from PENDING to PUBLISHED without a human, no examiner auto-merge, no key points before submission - are now checkable properties."),
    ("The engine needs no AI, and that is a measured claim rather than a preference.",
     "Cross-language semantic matching, negation-aware grading, typo tolerance and partial credit all work with zero model calls. Evaluation costs 0.36 ms and always returns the same answer for the same input."),
    ("Against commercial OSCE platforms it wins on provenance and loses on logistics.",
     "Speedwell and ExamSoft run exam-day circuits with tablet marking; this does not, and should not. Neither ingests historical recall material or resolves examiner identity from it, which is this system's actual asset."),
    ("Against the open-source assessment stack, the gap is interoperability.",
     "TAO is QTI-certified in all four categories. This engine exports nothing standard yet. The domain model is a superset of QTI's item model, so export is additive work, not a redesign."),
    ("Performance is not the risk; content quality is.",
     "Both exam-path budgets are consumed to under a fifth of one percent by engine compute. The unresolved question is extraction precision on real files, which needs a labelled corpus that does not exist yet."),
    ("The honest limitation is vocabulary coverage.",
     "A paraphrase naming no known concept is missed. Every such term is reported in unmatchedTerms, turning the ceiling into a reviewer work queue rather than a silent failure. That trade is deliberate."),
]
for title, body in findings:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(name="Calibri", size=10, bold=True, color=INK)
    c.alignment = Alignment(vertical="top", wrap_text=True)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value=body)
    c.font = Font(name="Calibri", size=10, color=MUTED)
    c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 32
    row += 2

for i, w in enumerate([22, 16, 8, 22, 16, 8, 22, 16], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ===========================================================================
# Sheet 2 - Framework coverage
# ===========================================================================
ws = wb.create_sheet("Framework coverage")
ws.sheet_view.showGridLines = False
r = sheet_title(
    ws, "Framework coverage",
    "Every numbered requirement of the source framework, its implementation status, and the artefact that evidences it.",
    5)
r = write_table(
    ws, r,
    ["Section", "Requirement", "Status", "Implementation", "Evidence"],
    [list(x) for x in D.COVERAGE],
    [10, 46, 22, 60, 44],
    status_cols=(3,), name="Coverage")
note(ws, r + 2,
     "Status values: Implemented = present and tested. Implemented + extended = present, plus capability beyond the framework's requirement. "
     "Verified by measurement = a performance requirement confirmed by the benchmark. Deferred, deliberately = not built, with the measured reason recorded.",
     5)

# ===========================================================================
# Sheet 3 - Competitive matrix
# ===========================================================================
ws = wb.create_sheet("Competitive matrix")
ws.sheet_view.showGridLines = False
names = [c["name"] for c in D.COMPETITORS]
r = sheet_title(
    ws, "Competitive capability matrix",
    "This engine against institutional OSCE platforms, open-source assessment stacks, question banks and the specialist tooling it borrows from. "
    "Full / Partial / None / N/A, where N/A means the capability is outside that product's category.",
    len(names) + 2)

headers = ["Capability"] + names + ["Reading"]
rows = [[cap] + ratings + [reading] for cap, ratings, reading in D.CAPABILITY_MATRIX]
r = write_table(
    ws, r, headers, rows,
    [46] + [15] * len(names) + [62],
    status_cols=tuple(range(2, 2 + len(names))), name="Capability")

r += 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(names) + 2)
c = ws.cell(row=r, column=1, value="Products compared")
c.font = Font(name="Calibri", size=12, bold=True, color=ACCENT)
r += 1
r = write_table(
    ws, r,
    ["Product", "Category", "Type", "What it is known for"],
    [[c["name"], c["category"], c["kind"], c["notes"]] for c in D.COMPETITORS],
    [26, 34, 16, 96], name="Products")

# ===========================================================================
# Sheet 4 - Deterministic vs LLM
# ===========================================================================
ws = wb.create_sheet("Deterministic vs LLM")
ws.sheet_view.showGridLines = False
r = sheet_title(
    ws, "Deterministic engine vs LLM in the path",
    "The decision record behind 'no AI needed'. Each row is a dimension on which the choice was actually made, including the one where the LLM wins.",
    4)
r = write_table(
    ws, r,
    ["Dimension", "Deterministic engine", "LLM in the path", "Which wins, and why"],
    [list(x) for x in D.DETERMINISM_MATRIX],
    [26, 52, 52, 56], name="Determinism")
note(ws, r + 2,
     "The LLM wins outright on paraphrase handling and on cold-start quality before any vocabulary exists. Both are real. "
     "The engine answers them by reporting every unrecognised term rather than hiding the miss, and by keeping the provider interfaces open so a "
     "semantic adapter can be added for the residue - but never on the critical path, and never as the system of record.",
     4)

# ===========================================================================
# Sheet 5 - Acceptance tests
# ===========================================================================
ws = wb.create_sheet("Acceptance tests")
ws.sheet_view.showGridLines = False
r = sheet_title(
    ws, "Acceptance tests",
    "The framework's Section 14 invariants, implemented as executable tests rather than prose.",
    5)
r = write_table(
    ws, r,
    ["#", "Test", "Expected invariant", "Result", "How it is verified"],
    [list(x) for x in D.ACCEPTANCE],
    [8, 32, 46, 14, 82],
    status_cols=(4,), name="Acceptance")

# ===========================================================================
# Sheet 6 - KPIs
# ===========================================================================
ws = wb.create_sheet("KPIs")
ws.sheet_view.showGridLines = False
r = sheet_title(
    ws, "Engineering KPIs",
    "The framework's Section 14 targets against measured values. Latency figures are engine CPU only; endpoint latency adds database and network time.",
    5)
r = write_table(
    ws, r,
    ["KPI", "Target", "Measured", "Status", "Note"],
    [list(x) for x in D.KPIS],
    [40, 20, 26, 14, 76],
    status_cols=(4,), name="KPIs")

r += 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
c = ws.cell(row=r, column=1, value="Benchmark detail")
c.font = Font(name="Calibri", size=12, bold=True, color=ACCENT)
r += 1
r = write_table(
    ws, r,
    ["Operation", "Corpus", "p50 (ms)", "p95 (ms)", "p99 (ms)", "Samples"],
    [list(x) for x in D.BENCHMARKS],
    [46, 42, 12, 12, 12, 12], name="Bench")
note(ws, r + 2,
     "The degenerate-blocking row is measured deliberately. It is the failure mode where every examiner name shares one phonetic skeleton, "
     "and it exists in the benchmark so the cost of that failure is known in advance rather than discovered in production.",
     6)

# ===========================================================================
# Sheet 7 - Risk register
# ===========================================================================
ws = wb.create_sheet("Risk register")
ws.sheet_view.showGridLines = False
r = sheet_title(
    ws, "Risk register",
    "Every risk in the framework's Section 11 table, plus the ones this implementation surfaced. Status is Closed only where a test enforces the control.",
    8)
r = write_table(
    ws, r,
    ["ID", "Risk", "Class", "Impact", "Likelihood", "Control", "Status", "Evidence"],
    [list(x) for x in D.RISKS],
    [7, 44, 18, 12, 13, 62, 20, 40],
    status_cols=(7,), name="Risks")

# ===========================================================================
# Sheet 8 - Roadmap
# ===========================================================================
ws = wb.create_sheet("Roadmap")
ws.sheet_view.showGridLines = False
r = sheet_title(
    ws, "Implementation roadmap",
    "The framework's Phase A-D plan, extended. Phases A-C are complete and measured; Phase D is blocked on data that does not exist yet, not on engineering.",
    5)
r = write_table(
    ws, r,
    ["Phase", "Work", "Status", "Detail", "Why it matters"],
    [list(x) for x in D.ROADMAP],
    [12, 34, 20, 74, 52],
    status_cols=(3,), name="Roadmap")
note(ws, r + 2,
     "Phase E.3 and E.4 are triggered rather than scheduled. E.3 fires when the largest examiner blocking bucket exceeds roughly 2,000 records, "
     "which ExaminerIndex.stats reports directly. E.4 fires only if Phase D.1 shows the vocabulary's residue is large enough to justify a semantic adapter.",
     5)

# ===========================================================================
# Sheet 9 - Evidence and sources
# ===========================================================================
ws = wb.create_sheet("Evidence and sources")
ws.sheet_view.showGridLines = False
r = sheet_title(
    ws, "Evidence index and sources",
    "Every claim in this workbook traces to a command that can be re-run, or to a published source.",
    4)
r = write_table(
    ws, r,
    ["Claim", "Value", "Where it comes from", "Command to reproduce"],
    [list(x) for x in D.EVIDENCE],
    [26, 62, 40, 30], mono_cols=(4,), name="Evidence")

r += 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
c = ws.cell(row=r, column=1, value="Module inventory")
c.font = Font(name="Calibri", size=12, bold=True, color=ACCENT)
r += 1
r = write_table(
    ws, r,
    ["Layer", "Modules", "Contents"],
    [list(x) for x in D.CODE_STATS],
    [26, 12, 90], name="Modules")

r += 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
c = ws.cell(row=r, column=1, value="External sources consulted")
c.font = Font(name="Calibri", size=12, bold=True, color=ACCENT)
r += 1
r = write_table(
    ws, r,
    ["Source", "URL"],
    [list(x) for x in D.SOURCES],
    [64, 90], mono_cols=(2,), name="Sources")

# ===========================================================================
# Sheet - Shipped code review
# ===========================================================================
ws = wb.create_sheet("Shipped code review", 1)
ws.sheet_view.showGridLines = False
r = sheet_title(
    ws, "Code review: the shipped implementation",
    f"Findings against {D.SHIPPED_META['package']} ({D.SHIPPED_META['files_reviewed']} files, read but never modified). "
    f"Stack: {D.SHIPPED_META['stack']}.",
    6)

ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
c = ws.cell(row=r, column=1, value="What the shipped code already gets right")
c.font = Font(name="Calibri", size=12, bold=True, color=ACCENT)
r += 1
r = write_table(
    ws, r, ["Strength", "Detail"],
    [list(x) for x in D.SHIPPED_STRENGTHS],
    [40, 118], name="Strengths")

r += 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
c = ws.cell(row=r, column=1, value="Findings")
c.font = Font(name="Calibri", size=12, bold=True, color=ACCENT)
r += 1
r = write_table(
    ws, r,
    ["ID", "Severity", "Area", "Finding", "Why it matters", "Fix"],
    [list(x) for x in D.SHIPPED_FINDINGS],
    [7, 12, 18, 62, 74, 56],
    status_cols=(2,), name="Findings")
note(ws, r + 2,
     "Severity reflects consequence, not effort. F1 and F2 are Critical because both change a student's recorded mark: "
     "F1 awards marks for answers a reviewer would penalise, and F2 lets a student retry a question after seeing the "
     "correct answer. Both fixes are small.",
     6)

# ===========================================================================
# Sheet - Evaluator comparison
# ===========================================================================
ws = wb.create_sheet("Evaluator comparison", 2)
ws.sheet_view.showGridLines = False
r = sheet_title(
    ws, "Evaluator head-to-head",
    "The shipped evaluator and the V2 evaluator run against identical inputs, scored against what a medical reviewer "
    "would mark. Reproduce with: node --experimental-strip-types docs/compare_evaluators.ts",
    6)
r = write_table(
    ws, r,
    ["Case", "Student answer", "Reviewer", "Shipped", "V2", "What happens"],
    [list(x) for x in D.EVALUATOR_COMPARISON],
    [30, 52, 13, 13, 13, 74],
    status_cols=(), name="EvalCmp")

r += 1
ws.cell(row=r, column=1, value="Agreement with a reviewer").font = Font(name="Calibri", size=11, bold=True, color=INK)
ws.cell(row=r, column=3, value=f"{D.EVALUATOR_SCORE['shipped']} / {D.EVALUATOR_SCORE['total']}").font = Font(name="Calibri", size=11, bold=True, color=BAD)
ws.cell(row=r, column=4, value=f"{D.EVALUATOR_SCORE['v2']} / {D.EVALUATOR_SCORE['total']}").font = Font(name="Calibri", size=11, bold=True, color=OK)
r += 2
note(ws, r,
     "Three of the nine shipped disagreements are in the dangerous direction - they award marks the student did not earn. "
     "The negation case is the clearest: a student who writes 'there is no evidence of deep vein thrombosis' receives full "
     "credit for the DVT key point.",
     6)

# ===========================================================================
# Sheet - Migration path
# ===========================================================================
ws = wb.create_sheet("Migration path", 3)
ws.sheet_view.showGridLines = False
r = sheet_title(
    ws, "Migration path",
    "How to move the shipped implementation onto the V2 engine, ordered by consequence per unit of effort. "
    "No step requires a rewrite; every migration is additive.",
    6)
r = write_table(
    ws, r,
    ["Step", "Change", "Size", "Migration", "How", "Effect"],
    [list(x) for x in D.MIGRATION_STEPS],
    [8, 30, 22, 24, 76, 62], name="Migration")
note(ws, r + 2,
     "M1 and M2 together are under 30 lines and close both Critical findings. They are independent of everything else "
     "and can ship before the deployment question is settled.",
     6)

# ===========================================================================
# Sheet - Applied and verified
# ===========================================================================
ws = wb.create_sheet("Applied and verified", 4)
ws.sheet_view.showGridLines = False
r = sheet_title(
    ws, "Applied to the shipped app, and verified",
    "Changes made to osce-app/ in this session, and what was actually run against them. "
    "No Cloudflare resource was created, listed or modified.",
    5)
r = write_table(
    ws, r, ["ID", "Change", "Where", "What it does", "Evidence"],
    [list(x) for x in D.APPLIED_FIXES],
    [8, 34, 40, 88, 62], name="Applied")

r += 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
c = ws.cell(row=r, column=1, value="Verification run")
c.font = Font(name="Calibri", size=12, bold=True, color=ACCENT)
r += 1
r = write_table(
    ws, r, ["Check", "Result", "Note"],
    [list(x) for x in D.VERIFICATION_RUN],
    [40, 34, 100], name="Verify")
note(ws, r + 2,
     "Cloudflare deployment was not attempted and could not be: the session had no Cloudflare credentials, "
     "no wrangler authentication and an unauthorized MCP connector, and the handoff brief records that the "
     "owner halted the cutover pending account verification. See osce-app/DEPLOYMENT_RUNBOOK.md for the "
     "remaining sequence.",
     5)

out = Path(__file__).parent / "OSCE_Engine_Analysis.xlsx"
wb.save(out)
print(f"wrote {out} ({out.stat().st_size / 1024:.0f} KB, {len(wb.sheetnames)} sheets)")
print("sheets:", ", ".join(wb.sheetnames))
